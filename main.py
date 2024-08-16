import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from tkcalendar import DateEntry
import pandas as pd
from random import shuffle
from datetime import datetime, timedelta
import sqlite3
from openpyxl.styles import Font, PatternFill
import numpy as np
import logging
import random
from collections import defaultdict

# Loglama ayarları
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')


class NobetAgent:
    def __init__(self, action_size, state_size):
        self.action_size = action_size
        self.state_size = state_size
        self.q_table = np.zeros((state_size, action_size))  # Q-table

    def choose_action(self, state):
        # Epsilon-greedy action selection
        if np.random.uniform(0, 1) < 0.1:  # exploration
            return np.random.choice(self.action_size)
        else:  # exploitation
            return np.argmax(self.q_table[state, :])

    def learn(self, state, action, reward, next_state):
        # Q-learning update
        predict = self.q_table[state, action]
        target = reward + 0.95 * np.max(self.q_table[next_state, :])
        self.q_table[state, action] += 0.01 * (target - predict)


def create_database(db_name):
    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS nobet_gecmisi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            isim TEXT NOT NULL,
            toplam_nobet INTEGER DEFAULT 0,
            hafta_sonu_nobet INTEGER DEFAULT 0
        )
    ''')
    conn.commit()
    return conn


def update_nobet_gecmisi(conn, isim, hafta_sonu=False):
    cursor = conn.cursor()
    if hafta_sonu:
        cursor.execute('''
            UPDATE nobet_gecmisi 
            SET toplam_nobet = toplam_nobet + 1, hafta_sonu_nobet = hafta_sonu_nobet + 1 
            WHERE isim = ?
        ''', (isim,))
    else:
        cursor.execute('''
            UPDATE nobet_gecmisi 
            SET toplam_nobet = toplam_nobet + 1 
            WHERE isim = ?
        ''', (isim,))

    # Eğer personelin kaydı yoksa, yeni bir kayıt oluştur
    if cursor.rowcount == 0:
        if hafta_sonu:
            cursor.execute('''
                INSERT INTO nobet_gecmisi (isim, toplam_nobet, hafta_sonu_nobet) 
                VALUES (?, 1, 1)
            ''', (isim,))
        else:
            cursor.execute('''
                INSERT INTO nobet_gecmisi (isim, toplam_nobet, hafta_sonu_nobet) 
                VALUES (?, 1, 0)
            ''', (isim,))

    conn.commit()


def get_nobet_gecmisi(conn, isim):
    cursor = conn.cursor()
    cursor.execute('SELECT toplam_nobet, hafta_sonu_nobet FROM nobet_gecmisi WHERE isim = ?', (isim,))
    result = cursor.fetchone()
    if result is None:
        cursor.execute('''
            INSERT INTO nobet_gecmisi (isim, toplam_nobet, hafta_sonu_nobet) 
            VALUES (?, 0, 0)
        ''', (isim,))
        conn.commit()
        return (0, 0)
    return result


def veritabani_baglantisi():
    conn = sqlite3.connect('nobet_app.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS personel (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            isim TEXT NOT NULL,
            grup INTEGER NOT NULL DEFAULT 1,
            haftasonu_nobeti INTEGER NOT NULL DEFAULT 1,
            izin_baslangic TEXT,
            izin_bitis TEXT
        )
    ''')
    conn.commit()
    return conn


def veritabani_baglantisi_haftasonu():
    conn = sqlite3.connect('haftasonu_nobet_gecmisi.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS haftasonu_nobet (
            isim TEXT PRIMARY KEY,
            haftasonu_nobet_sayisi INTEGER NOT NULL DEFAULT 0
        )
    ''')
    conn.commit()
    return conn


def personel_kaydet(conn, personel_listesi):
    cursor = conn.cursor()
    cursor.execute("DELETE FROM personel")
    for personel in personel_listesi:
        cursor.execute(
            "INSERT INTO personel (isim, grup, haftasonu_nobeti, izin_baslangic, izin_bitis) VALUES (?, ?, ?, ?, ?)",
            (personel["isim"], personel["grup"], int(personel["haftasonu_nobeti"]), personel.get("izin_baslangic"),
             personel.get("izin_bitis")))
    conn.commit()


def personel_yukle(conn):
    cursor = conn.cursor()
    cursor.execute("SELECT isim, grup, haftasonu_nobeti, izin_baslangic, izin_bitis FROM personel")
    rows = cursor.fetchall()
    personel_listesi = [{"isim": row[0], "grup": row[1], "haftasonu_nobeti": bool(row[2]), "izin_baslangic": row[3],
                         "izin_bitis": row[4]} for row in rows]
    return personel_listesi


def tarih_araligi_olustur(baslangic_tarihi, personel_sayisi):
    tarih_araligi = []
    baslangic = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
    bitis = baslangic + timedelta(days=personel_sayisi - 1)
    gün = baslangic

    while gün <= bitis:
        tarih_araligi.append(gün.strftime('%Y-%m-%d'))
        gün += timedelta(days=1)

    return tarih_araligi


def hafta_ici_ve_hafta_sonu_ayir(tarih_araligi):
    hafta_ici_gunleri = []
    hafta_sonu_gunleri = []

    for tarih in tarih_araligi:
        gün = datetime.strptime(tarih, '%Y-%m-%d').weekday()
        if gün < 5:  # Pazartesi=0, Cuma=4, Cumartesi=5, Pazar=6
            hafta_ici_gunleri.append(tarih)
        else:
            hafta_sonu_gunleri.append(tarih)

    return hafta_ici_gunleri, hafta_sonu_gunleri


def haftasonu_nobet_guncelle(conn, personel_isim):
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO haftasonu_nobet (isim, haftasonu_nobet_sayisi)
        VALUES (?, 1)
        ON CONFLICT(isim) DO UPDATE SET haftasonu_nobet_sayisi = haftasonu_nobet_sayisi + 1
    ''', (personel_isim,))
    conn.commit()


def haftasonu_nobet_sayisi_al(conn, personel_isim):
    cursor = conn.cursor()
    cursor.execute('SELECT haftasonu_nobet_sayisi FROM haftasonu_nobet WHERE isim = ?', (personel_isim,))
    row = cursor.fetchone()
    if row:
        return row[0]
    return 0


def izinli_mi(personel, tarih):
    izin_baslangic = personel.get("izin_baslangic")
    izin_bitis = personel.get("izin_bitis")
    if izin_baslangic and izin_bitis:
        izin_baslangic = datetime.strptime(izin_baslangic, '%Y-%m-%d')
        izin_bitis = datetime.strptime(izin_bitis, '%Y-%m-%d')
        tarih_obj = datetime.strptime(tarih, '%Y-%m-%d')
        if izin_baslangic <= tarih_obj <= izin_bitis:
            return True
    return False


def denge_nobet_cizelgesi(nobet_cizelgesi, personel_listesi, haftasonu_conn):
    # Nöbet sayısı ve hafta sonu nöbet sayısını dengelemek için ek bir kontrol fonksiyonu
    for _ in range(5):  # 5 kez dengeleme denemesi yapalım
        toplam_nobetler = {personel["isim"]: get_nobet_gecmisi(haftasonu_conn, personel["isim"])[0] for personel in
                           personel_listesi}
        hafta_sonu_nobetler = {personel["isim"]: get_nobet_gecmisi(haftasonu_conn, personel["isim"])[1] for personel in
                               personel_listesi}
        # En az ve en çok nöbet tutanları belirleyelim
        en_az_nobet = min(toplam_nobetler.values())
        en_cok_nobet = max(toplam_nobetler.values())

        # En az nöbet tutanlardan birini ve en çok nöbet tutanlardan birini seçelim
        az_nobet_personelleri = [isim for isim, sayi in toplam_nobetler.items() if sayi == en_az_nobet]
        cok_nobet_personelleri = [isim for isim, sayi in toplam_nobetler.items() if sayi == en_cok_nobet]

        # Aynı şekilde hafta sonu nöbetleri için de kontrol yapalım
        en_az_hafta_sonu = min(hafta_sonu_nobetler.values())
        en_cok_hafta_sonu = max(hafta_sonu_nobetler.values())

        az_hafta_sonu_personelleri = [isim for isim, sayi in hafta_sonu_nobetler.items() if sayi == en_az_hafta_sonu]
        cok_hafta_sonu_personelleri = [isim for isim, sayi in hafta_sonu_nobetler.items() if sayi == en_cok_hafta_sonu]

        # Nöbet günlerini dengelemek için nöbetçilerin yerlerini değiştirelim
        for tarih, nobetciler in nobet_cizelgesi.items():
            for i in range(len(nobetciler)):
                if nobetciler[i] in cok_nobet_personelleri:
                    uygun_personel = next((p for p in az_nobet_personelleri if p not in nobetciler), None)
                    if uygun_personel:
                        nobetciler[i] = uygun_personel

            # Hafta sonu nöbetlerini de dengeleyelim
            if datetime.strptime(tarih, '%Y-%m-%d').weekday() >= 5:  # Hafta sonu kontrolü
                for i in range(len(nobetciler)):
                    if nobetciler[i] in cok_hafta_sonu_personelleri:
                        uygun_personel = next((p for p in az_hafta_sonu_personelleri if p not in nobetciler), None)
                        if uygun_personel:
                            nobetciler[i] = uygun_personel

    return nobet_cizelgesi


def nöbet_çizelgesi_oluştur(personel_listesi, izinli_listesi, tarih_araligi, max_nobet_sayisi, min_gun_araligi,
                            haftasonu_conn):
    def uygun_mu(nobet_çizelgesi, personel, gün):
        # Son nöbet günü kontrolü
        son_nobet_gunu = son_nöbet_günleri[personel["isim"]]
        if son_nobet_gunu is not None:
            gun_farki = (datetime.strptime(gün, '%Y-%m-%d') - datetime.strptime(son_nobet_gunu, '%Y-%m-%d')).days
            if gun_farki < min_gun_araligi:
                return False

        # Maksimum nöbet sayısı kontrolü
        if nöbet_sayilari[personel["isim"]] >= max_nobet_sayisi:
            return False

        # Hafta sonu kontrolü
        hafta_sonu_gunu = datetime.strptime(gün, '%Y-%m-%d').weekday() >= 5
        if hafta_sonu_gunu and not personel["haftasonu_nobeti"]:
            return False

        # İzin kontrolü
        if izinli_mi(personel, gün):
            return False

        return True

    def çizelgeyi_kontrol_et(nobet_çizelgesi):
        # Boş gün var mı?
        for nobetçiler in nöbet_çizelgesi.values():
            if len(nobetçiler) < 2:
                return False

        # Nöbet sayıları adil mi?
        for personel, sayi in nöbet_sayilari.items():
            if sayi < max_nobet_sayisi:
                return False

        return True

    # Başlangıç değerleri
    nöbet_çizelgesi = defaultdict(list)
    nöbet_sayilari = defaultdict(int)
    son_nöbet_günleri = defaultdict(lambda: None)
    hafta_sonu_nobet_sayisi = defaultdict(int)

    deneme = 0
    while True:
        deneme += 1
        logging.debug(f"{deneme}. deneme: Nöbet çizelgesi oluşturuluyor...")

        # Yeni bir çizelge oluştur
        nöbet_çizelgesi.clear()
        nöbet_sayilari.clear()
        son_nöbet_günleri.clear()
        hafta_sonu_nobet_sayisi.clear()

        # Grupları sıraya göre yerleştir
        gruplar = defaultdict(list)
        for personel in personel_listesi:
            gruplar[personel["grup"]].append(personel)

        for gün in tarih_araligi:
            uygun_personeller = []
            for grup, grup_uyeleri in gruplar.items():
                random.shuffle(grup_uyeleri)
                uygun_personeller.extend([p for p in grup_uyeleri if uygun_mu(nöbet_çizelgesi, p, gün)])

            random.shuffle(uygun_personeller)

            # Önce aynı gruptan olanları atamaya çalış
            for personel in uygun_personeller:
                if len(nöbet_çizelgesi[gün]) < 2:
                    grup_uyesi = any(p["grup"] == personel["grup"] for p in nöbet_çizelgesi[gün])
                    if grup_uyesi or not nöbet_çizelgesi[gün]:
                        nöbet_çizelgesi[gün].append(personel)
                        nöbet_sayilari[personel["isim"]] += 1
                        son_nöbet_günleri[personel["isim"]] = gün
                        if datetime.strptime(gün, '%Y-%m-%d').weekday() >= 5:
                            hafta_sonu_nobet_sayisi[personel["isim"]] += 1

            # Eğer aynı gruptan uygun kimse kalmadıysa diğer gruplardan alınabilir
            for personel in uygun_personeller:
                if len(nöbet_çizelgesi[gün]) < 2:
                    if not any(p["grup"] == personel["grup"] for p in nöbet_çizelgesi[gün]):
                        nöbet_çizelgesi[gün].append(personel)
                        nöbet_sayilari[personel["isim"]] += 1
                        son_nöbet_günleri[personel["isim"]] = gün
                        if datetime.strptime(gün, '%Y-%m-%d').weekday() >= 5:
                            hafta_sonu_nobet_sayisi[personel["isim"]] += 1

        # Eğer nöbetçi atanamayan günler varsa, hafta sonu durumu olmayan personel atanabilir
        for gün in tarih_araligi:
            if len(nöbet_çizelgesi[gün]) < 2:
                for grup, grup_uyeleri in gruplar.items():
                    for personel in grup_uyeleri:
                        if personel["haftasonu_nobeti"] == False and uygun_mu(nöbet_çizelgesi, personel, gün):
                            nöbet_çizelgesi[gün].append(personel)
                            nöbet_sayilari[personel["isim"]] += 1
                            son_nöbet_günleri[personel["isim"]] = gün
                            if datetime.strptime(gün, '%Y-%m-%d').weekday() >= 5:
                                hafta_sonu_nobet_sayisi[personel["isim"]] += 1
                        if len(nöbet_çizelgesi[gün]) >= 2:
                            break

        # Çizelgeyi kontrol et
        if çizelgeyi_kontrol_et(nöbet_çizelgesi):
            logging.debug(f"{deneme}. deneme: Başarılı!")
            # Başarılı olunduğunda nöbet geçmişini güncelle
            for personel in personel_listesi:
                if nöbet_sayilari[personel["isim"]] > 0:  # Sadece nöbet yazılanları güncelle
                    update_nobet_gecmisi(haftasonu_conn, personel["isim"],
                                         hafta_sonu_nobet_sayisi[personel["isim"]] > 0)
            break
        else:
            logging.debug(f"{deneme}. deneme: Kurallara uyulmadı, tekrar deneniyor...")

    return nöbet_çizelgesi, nöbet_sayilari, hafta_sonu_nobet_sayisi



def excel_yaz(nobet_çizelgesi, dosya_adi, personel_listesi, toplam_nobet_sayisi, hafta_sonu_nobet_sayisi,
              haftasonu_conn):
    try:
        with pd.ExcelWriter(dosya_adi, engine='openpyxl') as writer:
            if nobet_çizelgesi:  # Nöbetler boş değilse
                # Nöbetçiler listesini eşitlemek için
                for gün in nobet_çizelgesi:
                    while len(nobet_çizelgesi[gün]) < 2:
                        nobet_çizelgesi[gün].append(None)

                # Sadece isimleri almak için güncelleme yapılıyor
                df = pd.DataFrame.from_dict(
                    {gün: [nobetci.get('isim') if nobetci else None for nobetci in nobet_çizelgesi[gün]]
                     for gün in nobet_çizelgesi},
                    orient='index', columns=["Nöbetçi 1", "Nöbetçi 2"]
                )

                df.index.name = 'Tarih'
                df.reset_index(inplace=True)

                # Gün isimlerini ekleyelim
                df['Gün'] = df['Tarih'].apply(lambda x: datetime.strptime(x, '%Y-%m-%d').strftime('%A'))

                # Excel'e yazma
                df.to_excel(writer, sheet_name="Nöbet Çizelgesi", index=False)

                # Açılan Excel dosyasına biçimlendirme uygulamak için erişim sağlıyoruz
                workbook = writer.book
                sheet = workbook["Nöbet Çizelgesi"]

                # Hafta sonu günlerini kırmızı renkle işaretleme
                red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

                for row in range(2, len(df) + 2):  # 2, çünkü Excel'de başlık satırı 1. satır
                    gün_hücre = sheet.cell(row=row, column=4)  # 'Gün' sütunu Excel'de 4. sütun
                    gün_isim = gün_hücre.value
                    if gün_isim in ['Saturday', 'Sunday']:  # Cumartesi ve Pazar günlerini kontrol et
                        for col in range(1, 5):  # Tarih, Nöbetçi 1, Nöbetçi 2 ve Gün sütunlarını renklendirme
                            sheet.cell(row=row, column=col).fill = red_fill

                    # Nöbetçilerin sarı renkle işaretlenmesi
                    for col in range(2, 4):  # Nöbetçi 1 ve Nöbetçi 2 sütunları
                        personel_hücre = sheet.cell(row=row, column=col)
                        if personel_hücre.value:
                            nöbet_sayisi = sum(df['Nöbetçi 1'].eq(personel_hücre.value)) + sum(
                                df['Nöbetçi 2'].eq(personel_hücre.value))
                            if nöbet_sayisi == 2:
                                personel_hücre.fill = yellow_fill

                # Personel Listesi ve Nöbet Geçmişi Sayfası
                personel_data = []
                for personel in personel_listesi:
                    toplam_nobet, hafta_sonu_nobet = get_nobet_gecmisi(haftasonu_conn, personel["isim"])
                    toplam_nobet += toplam_nobet_sayisi.get(personel["isim"], 0)  # Güncel toplam nöbet sayısını ekle
                    hafta_sonu_nobet += hafta_sonu_nobet_sayisi.get(personel["isim"],
                                                                    0)  # Güncel hafta sonu nöbet sayısını ekle

                    personel_data.append({
                        "İsim": personel["isim"],
                        "Grup": personel["grup"],
                        "Toplam Nöbet": toplam_nobet,
                        "Hafta Sonu Nöbet": hafta_sonu_nobet
                    })

                personel_df = pd.DataFrame(personel_data)
                personel_df.to_excel(writer, sheet_name="Nöbet Geçmişi", index=False)

    except Exception as e:
        messagebox.showerror("Hata", f"Excel dosyası oluşturulurken bir hata oluştu: {str(e)}")


class NobetApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Nöbet Yönetim Sistemi")
        self.geometry("600x1000")

        self.conn = veritabani_baglantisi()

        self.personel_listesi = personel_yukle(self.conn)
        self.izinli_listesi = {}

        self.setup_widgets()

    def setup_widgets(self):
        # Personel ekleme formu
        self.personel_frame = tk.Frame(self)
        self.personel_frame.grid(row=0, column=0, padx=5, pady=5, sticky="w")

        self.personel_entry = tk.Entry(self.personel_frame, width=40, )
        self.personel_entry.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.personel_entry.insert(0, "İsim Giriniz")
        self.personel_entry.bind("<FocusIn>", self.on_entry_click)
        self.personel_entry.bind("<FocusOut>", self.on_focusout)

        self.grup_var = tk.IntVar()
        self.grup_label = tk.Label(self.personel_frame, text="Grup:")
        self.grup_label.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        self.grup_spinbox = tk.Spinbox(self.personel_frame, from_=1, to=10, textvariable=self.grup_var, width=3)
        self.grup_spinbox.grid(row=0, column=2, padx=5, pady=5, sticky="w")

        self.ekle_button = tk.Button(self.personel_frame, text="Personel Ekle", command=self.ekle_personel,
                                     fg="white",
                                     bg="green")
        self.ekle_button.grid(row=0, column=3, padx=5, pady=5, sticky="w")

        # Personel Listesi Görüntüleme (Tablo olarak Treeview kullanıldı)
        self.personel_tree = ttk.Treeview(self, columns=("isim", "grup", "haftasonu", "izin"), show='headings',
                                          height=28)
        self.personel_tree.grid(row=1, column=0, columnspan=4, padx=10, pady=10, sticky="nsew")

        self.personel_tree.heading("isim", text="Kişi")
        self.personel_tree.heading("grup", text="Grup")
        self.personel_tree.heading("haftasonu", text="Hafta Sonu Durumu")
        self.personel_tree.heading("izin", text="İzin Durumu")

        self.personel_tree.column("isim", width=200)
        self.personel_tree.column("grup", width=30, anchor="center")
        self.personel_tree.column("haftasonu", width=110, anchor="center")
        self.personel_tree.column("izin", width=200, anchor="center")

        self.ayar_frame = tk.Frame(self, highlightbackground="red", highlightthickness=2, borderwidth=1)
        self.ayar_frame.grid(row=2, column=0, padx=5, pady=5, sticky="w")

        self.haftasonu_toggle_button = tk.Button(self.ayar_frame, text="Hafta Sonu Durumunu Değiştir",
                                                 command=self.haftasonu_toggle, fg="white",
                                                 bg="orange")
        self.haftasonu_toggle_button.grid(row=2, column=0, padx=5, pady=5, sticky="w")

        self.isim_duzenle_button = tk.Button(self.ayar_frame, text="İsim Düzenle", command=self.isim_duzenle,
                                             fg="white", bg="blue")
        self.isim_duzenle_button.grid(row=2, column=1, padx=10, pady=10, sticky="w")

        self.cikar_button = tk.Button(self.ayar_frame, text="Seçili Personeli Çıkar",
                                      command=self.cikar_personel,
                                      fg="white", bg="red")
        self.cikar_button.grid(row=2, column=2, padx=5, pady=5, sticky="w")

        self.guncelle_button = tk.Button(self.ayar_frame, text="Personel Listesini Güncelle",
                                         command=self.guncelle_personel_listesi, fg="white", bg="green")
        self.guncelle_button.grid(row=2, column=3, padx=5, pady=5, sticky="w")

        self.grup_degis_label = tk.Label(self.ayar_frame, text="Yeni Grup:")
        self.grup_degis_label.grid(row=4, column=0, padx=5, pady=5, sticky="w")

        self.grup_degis_spinbox = tk.Spinbox(self.ayar_frame, from_=1, to=10, width=3)
        self.grup_degis_spinbox.grid(row=4, column=1, padx=5, pady=5, sticky="w")

        self.grup_degistir_button = tk.Button(self.ayar_frame, text="Grubu Değiştir",
                                              command=self.grup_degistir,
                                              fg="white", bg="blue")
        self.grup_degistir_button.grid(row=4, column=2, padx=5, pady=5, sticky="w")

        self.izin_baslangic_label = tk.Label(self.ayar_frame, text="İzin Başlangıç Tarihi:")
        self.izin_baslangic_label.grid(row=5, column=0, padx=5, pady=5, sticky="w")

        self.izin_baslangic_entry = DateEntry(self.ayar_frame, width=12, background='darkblue',
                                              foreground='white', borderwidth=2, year=datetime.now().year)
        self.izin_baslangic_entry.grid(row=5, column=1, padx=5, pady=5, sticky="w")

        self.izin_bitis_label = tk.Label(self.ayar_frame, text="İzin Bitiş Tarihi:")
        self.izin_bitis_label.grid(row=6, column=0, padx=5, pady=5, sticky="w")

        self.izin_bitis_entry = DateEntry(self.ayar_frame, width=12, background='darkblue',
                                          foreground='white', borderwidth=2, year=datetime.now().year)
        self.izin_bitis_entry.grid(row=6, column=1, padx=5, pady=5, sticky="w")

        self.izin_ekle_button = tk.Button(self.ayar_frame, text="İzin Tarihi Ekle",
                                          command=self.izin_tarihi_ekle,
                                          fg="white", bg="orange")
        self.izin_ekle_button.grid(row=5, column=2, padx=5, pady=5, sticky="w")

        self.izin_sil_button = tk.Button(self.ayar_frame, text="İzin Tarihi Sil", command=self.izin_tarihi_sil,
                                         fg="white", bg="red")
        self.izin_sil_button.grid(row=6, column=2, padx=5, pady=5, sticky="w")

        self.nobet_frame = tk.Frame(self, highlightbackground="green", highlightthickness=2, borderwidth=1)
        self.nobet_frame.grid(row=9, column=0, padx=5, pady=5, sticky="w")

        self.min_gun_araligi_label = tk.Label(self.nobet_frame, text="Nöbetler Arası Minimum Gün Sayısı:")
        self.min_gun_araligi_label.grid(row=9, column=0, padx=5, pady=5, sticky="w")

        self.min_gun_araligi_var = tk.IntVar(value=10)  # Varsayılan değer 10 gün
        self.min_gun_araligi_spinbox = tk.Spinbox(self.nobet_frame, from_=1, to=30,
                                                  textvariable=self.min_gun_araligi_var, width=3)
        self.min_gun_araligi_spinbox.grid(row=9, column=1, padx=5, pady=5, sticky="w")

        self.baslangic_tarihi_label = tk.Label(self.nobet_frame, text="Başlangıç Tarihi (YYYY-AA-GG):")
        self.baslangic_tarihi_label.grid(row=10, column=0, padx=5, pady=5, sticky="w")

        yarin = datetime.now() + timedelta(days=1)
        self.baslangic_tarihi_entry = tk.Entry(self.nobet_frame)
        self.baslangic_tarihi_entry.insert(0, yarin.strftime('%Y-%m-%d'))
        self.baslangic_tarihi_entry.grid(row=10, column=1, padx=5, pady=5, sticky="w")

        self.nobet_sayisi_label = tk.Label(self.nobet_frame, text="Her Personel için Nöbet Sayısı:")
        self.nobet_sayisi_label.grid(row=11, column=0, padx=5, pady=5, sticky="w")

        self.nobet_sayisi_entry = tk.Entry(self.nobet_frame)
        self.nobet_sayisi_entry.insert(0, '2')
        self.nobet_sayisi_entry.grid(row=11, column=1, padx=5, pady=5, sticky="w")

        self.olustur_button = tk.Button(self.nobet_frame, text="Nöbet Çizelgesi Oluştur", command=self.olustur_cizelge,
                                        fg="white", bg="green")
        self.olustur_button.grid(row=12, column=0, columnspan=3, padx=5, pady=5, sticky="nsew")

        self.personel_listesini_goster()

    # Placeholder için gerekli fonksiyonlar
    def on_entry_click(self, event):
        if self.personel_entry.get() == 'İsim Giriniz':
            self.personel_entry.delete(0, "end")
            self.personel_entry.config(fg='black')

    def on_focusout(self, event):
        if self.personel_entry.get() == '':
            self.personel_entry.insert(0, 'İsim Giriniz')
            self.personel_entry.config(fg='grey')

    def ekle_personel(self):
        personel = self.personel_entry.get()
        grup = self.grup_var.get()
        if personel and 1 <= grup <= 10:
            self.personel_listesi.append({"isim": personel, "grup": grup, "haftasonu_nobeti": True})
            self.personel_listesini_goster()  # Listeyi güncelle
            messagebox.showinfo("Başarılı", f"{personel} eklendi.")
        else:
            messagebox.showerror("Hata", "Lütfen geçerli bir isim ve grup numarası girin.")
        self.personel_entry.delete(0, tk.END)

    def cikar_personel(self):
        selected_item = self.personel_tree.selection()  # Seçili olan Treeview öğesini al
        if selected_item:
            values = self.personel_tree.item(selected_item, "values")  # Seçili öğenin değerlerini al
            personel_adı = values[0]  # İlk sütun (isim sütunu)

            # Personel listesinden ilgili kişiyi bulup silme
            self.personel_listesi = [personel for personel in self.personel_listesi if personel["isim"] != personel_adı]
            res = messagebox.askquestion('Personel Sil', 'Silmek İstediğinizden Emin Misiniz?')
            if res == 'yes':
                self.personel_tree.delete(selected_item)  # Seçili öğeyi Treeview'dan sil
                messagebox.showinfo("Başarılı", f"{personel_adı} çıkarıldı.")
            else:
                messagebox.showinfo("Silinmedi", f"{personel_adı} Silinmedi.")
        else:
            messagebox.showerror("Hata", "Lütfen tablodan bir personel seçin.")

    def isim_duzenle(self):
        selected_item = self.personel_tree.selection()
        if selected_item:
            values = self.personel_tree.item(selected_item, "values")
            eski_isim = values[0]

            yeni_isim = simpledialog.askstring("İsim Düzenle", f"{eski_isim} için yeni ismi giriniz:")

            if yeni_isim:
                for personel in self.personel_listesi:
                    if personel["isim"] == eski_isim:
                        personel["isim"] = yeni_isim
                        break

                self.personel_tree.item(selected_item, values=(yeni_isim, personel["grup"], values[2], values[3]))
                messagebox.showinfo("Başarılı", f"{eski_isim} ismi {yeni_isim} olarak değiştirildi.")
        else:
            messagebox.showerror("Hata", "Lütfen tablodan bir personel seçin.")

    def haftasonu_toggle(self):
        selected_item = self.personel_tree.selection()
        if selected_item:
            values = self.personel_tree.item(selected_item, "values")
            personel_adı = values[0]

            for personel in self.personel_listesi:
                if personel["isim"] == personel_adı:
                    personel["haftasonu_nobeti"] = not personel["haftasonu_nobeti"]
                    break

            haftasonu_durum = "Var" if personel["haftasonu_nobeti"] else "Yok"
            self.personel_tree.item(selected_item,
                                    values=(personel["isim"], personel["grup"], haftasonu_durum, values[3]))

            messagebox.showinfo("Başarılı", f"{personel_adı} için hafta sonu nöbet durumu değiştirildi.")
        else:
            messagebox.showerror("Hata", "Lütfen tablodan bir personel seçin.")

    def grup_degistir(self):
        selected_item = self.personel_tree.selection()
        if selected_item:
            values = self.personel_tree.item(selected_item, "values")
            personel_adı = values[0]
            yeni_grup = int(self.grup_degis_spinbox.get())

            for personel in self.personel_listesi:
                if personel["isim"] == personel_adı:
                    personel["grup"] = yeni_grup
                    break

            self.personel_tree.item(selected_item, values=(personel["isim"], personel["grup"], values[2], values[3]))

            messagebox.showinfo("Başarılı", f"{personel_adı} için grup {yeni_grup} olarak değiştirildi.")
        else:
            messagebox.showerror("Hata", "Lütfen tablodan bir personel seçin.")

    def izin_tarihi_ekle(self):
        selected_item = self.personel_tree.selection()
        if selected_item:
            values = self.personel_tree.item(selected_item, "values")
            personel_adı = values[0]
            izin_baslangic = self.izin_baslangic_entry.get_date().strftime('%Y-%m-%d')
            izin_bitis = self.izin_bitis_entry.get_date().strftime('%Y-%m-%d')

            for personel in self.personel_listesi:
                if personel["isim"] == personel_adı:
                    personel["izin_baslangic"] = izin_baslangic
                    personel["izin_bitis"] = izin_bitis
                    break

            izin_durum = f"{izin_baslangic} - {izin_bitis}"
            self.personel_tree.item(selected_item, values=(personel["isim"], personel["grup"], values[2], izin_durum))

            messagebox.showinfo("Başarılı", f"{personel_adı} için izin tarihleri eklendi.")
        else:
            messagebox.showerror("Hata", "Lütfen tablodan bir personel seçin.")

    def izin_tarihi_sil(self):
        selected_item = self.personel_tree.selection()
        if selected_item:
            values = self.personel_tree.item(selected_item, "values")
            personel_adı = values[0]

            for personel in self.personel_listesi:
                if personel["isim"] == personel_adı:
                    personel["izin_baslangic"] = None
                    personel["izin_bitis"] = None
                    break

            izin_durum = ""
            self.personel_tree.item(selected_item, values=(personel["isim"], personel["grup"], values[2], izin_durum))

            messagebox.showinfo("Başarılı", f"{personel_adı} için izin tarihleri silindi.")
        else:
            messagebox.showerror("Hata", "Lütfen tablodan bir personel seçin.")

    def guncelle_personel_listesi(self):
        personel_kaydet(self.conn, self.personel_listesi)
        messagebox.showinfo("Başarılı", "Personel listesi güncellendi ve kaydedildi.")

    def personel_listesini_goster(self):
        # Mevcut verileri temizleyelim
        for row in self.personel_tree.get_children():
            self.personel_tree.delete(row)

        for personel in self.personel_listesi:
            haftasonu_durum = "Var" if personel["haftasonu_nobeti"] else "Yok"
            izin_durum = f"{personel.get('izin_baslangic', 'Yok')} - {personel.get('izin_bitis', 'Yok')}"
            self.personel_tree.insert("", "end",
                                      values=(personel["isim"], personel["grup"], haftasonu_durum, izin_durum))

    def olustur_cizelge(self):
        logging.debug("olustur_cizelge fonksiyonu çalışıyor")
        baslangic_tarihi = self.baslangic_tarihi_entry.get()
        try:
            datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
        except ValueError:
            messagebox.showerror("Hata", "Geçersiz tarih formatı, lütfen YYYY-AA-GG formatında girin.")
            return

        try:
            max_nobet_sayisi = int(self.nobet_sayisi_entry.get())
        except ValueError:
            messagebox.showerror("Hata", "Nöbet sayısı için geçerli bir sayı girin.")
            return

        min_gun_araligi = self.min_gun_araligi_var.get()

        # Haftasonu nöbet geçmişi veritabanı bağlantısı
        haftasonu_conn = create_database('nobet_app2.db')

        logging.debug("nöbet_çizelgesi_oluştur fonksiyonu çağrılacak")
        # Nöbet Çizelgesini Oluştur
        tarih_araligi = tarih_araligi_olustur(baslangic_tarihi, len(self.personel_listesi))
        nobet_cizelgesi, toplam_nobet_sayisi, hafta_sonu_nobet_sayisi = nöbet_çizelgesi_oluştur(
            self.personel_listesi,
            self.izinli_listesi,
            tarih_araligi,
            max_nobet_sayisi,
            min_gun_araligi,
            haftasonu_conn
        )

        logging.debug("nöbet_çizelgesi_oluştur fonksiyonu çağrıldı")

        # Sonucu Dosyaya Yaz
        excel_yaz(nobet_cizelgesi, 'nobet_cizelgesi.xlsx', self.personel_listesi, toplam_nobet_sayisi,
                  hafta_sonu_nobet_sayisi, haftasonu_conn)  # haftasonu_conn parametresini ekleyin
        messagebox.showinfo("Başarılı", "Nöbet çizelgesi oluşturuldu ve 'nobet_cizelgesi.xlsx' dosyasına kaydedildi.")


if __name__ == "__main__":
    app = NobetApp()
    app.mainloop()
